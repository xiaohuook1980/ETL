"""dispatcher：文件流唯一编排者（task #13）

逻辑：
  for fid in raw_files.pending:
      file = load_xlsx(fid)
      result = identifier.identify(file)
      target_path = 归属判断 → 项目层/企业层/实控人层 + "考账发工" 子目录
      cos_ops.relocate(src, target_path)
      raw_files.ai_cos_key = target_path
      raw_files.detected_type = result.file_detected_type
      for sheet in result.sheets:
          parser_for_kind[sheet.kind].process_sheet(...)
      raw_files.parse_status = 'parsed' if 全部成功 else 'failed'

事务边界：整文件级。任一 parser 失败 → 整文件 rollback + parse_status='failed'。

归属判断规则：
  source_project_ids 数组：
    1 个项目                            → 项目层（企业/{ent_short}/项目/{proj_short}/考账发工/）
    多项目同企业                        → 企业层（企业/{ent_short}/考账发工/）
    多项目跨企业（同实控人）            → 实控人层（实控人/{name}/考账发工/）
    多项目跨实控人（不应发生 → 记日志） → 默认项目层取首项 + log warn
"""
import os
import sys
import io
import json
import argparse
sys.path.insert(0, 'D:/小鱼AI数据')

# by-handler 路由：默认 ON（4 kind 全开）
# 关闭走老 KIND_TO_PARSER：DISPATCH_HANDLER_ROUTE=0 python ...
USE_HANDLER_ROUTE = os.getenv('DISPATCH_HANDLER_ROUTE', '1').lower() in ('1', 'true', 'yes')
from openpyxl import load_workbook
from scripts._db import connect
from scripts._cos import download_bytes
from etl.classify import detect_mime
from etl.identifier import identify
from etl import cos_ops
from etl.parsers import (parse_attendance, parse_wage,
                          parse_bill, parse_payroll_xlsx,
                          parse_attendance_pdf, parse_bill_pdf)


KIND_TO_PARSER = {
    'attendance': parse_attendance,
    'wage_sheet': parse_wage,
    'bill': parse_bill,
    'payroll': parse_payroll_xlsx,
}


def _classify_pdf_by_filename(filename):
    """简单按文件名识别 PDF 类型。返回 'attendance' / 'bill' / 'unknown'"""
    if not filename:
        return 'unknown'
    name = filename
    if '考勤' in name:
        return 'attendance'
    if '计费明细' in name:
        return 'bill'
    return 'unknown'


def _load_xls_as_openpyxl(content, filename=None):
    """旧 .xls (BIFF8) → 用 xlrd 读 → 模拟 openpyxl Workbook
    parsers 当前用 openpyxl ws.iter_rows 接口；这里在内存重建一个兼容的 Workbook

    加密 xls (OLE2 加密)：先尝试 _config/密码.txt 精确匹配 + 文件名 regex 提取 → 解密 → 再读。
    解密后内容可能是 xlsx（PK 头）或 xls（BIFF8）。
    匹中失败 → 抛 RuntimeError 明确报错。"""
    import xlrd
    from openpyxl import Workbook
    try:
        xb = xlrd.open_workbook(file_contents=content)
    except xlrd.XLRDError as e:
        # OLE2 加密 → 尝试解密
        if 'OLE2' in str(e) or 'encrypted' in str(e).lower() or 'compound' in str(e).lower():
            from etl.raw.passwords import lookup_password, decrypt_office
            pwd = lookup_password(filename or '')
            if not pwd:
                raise RuntimeError(
                    f'加密 xls 但找不到密码。文件: {filename!r}。'
                    f'请在 ai 桶 `_config/密码.txt` 加一行 `{filename}<TAB><密码>`，'
                    f'或文件名加"密码XXX"提示（如 "xx-密码abc.xls"）'
                )
            decrypted = decrypt_office(content, pwd)
            # 解密后可能是 xlsx (PK 头) → 走 openpyxl；否则继续 xlrd
            if decrypted[:2] == b'PK':
                return load_workbook(io.BytesIO(decrypted), data_only=True)
            xb = xlrd.open_workbook(file_contents=decrypted)
        else:
            raise
    wb = Workbook()
    wb.remove(wb.active)
    for xs in xb.sheets():
        # xlrd: visibility 0=visible / 1=hidden / 2=very hidden  → 仅 0 保留
        if getattr(xs, 'visibility', 0) != 0:
            continue
        ws = wb.create_sheet(xs.name[:31])  # openpyxl sheet name 限 31 字符
        for ri in range(xs.nrows):
            row = []
            for ci in range(xs.ncols):
                v = xs.cell_value(ri, ci)
                # xlrd 日期是 float，转 datetime
                ct = xs.cell_type(ri, ci)
                if ct == xlrd.XL_CELL_DATE:
                    try:
                        from datetime import datetime
                        v = xlrd.xldate_as_datetime(v, xb.datemode)
                    except Exception:
                        pass
                elif ct == xlrd.XL_CELL_EMPTY:
                    v = None
                row.append(v if v != '' else None)
            ws.append(row)
    return wb


def _load_xlsx_via_calamine(content):
    """xlsx 经 openpyxl 失败时（如 'could not read stylesheet from None' 损坏的样式表）
    fallback 用 python-calamine 读取，重建为 openpyxl Workbook 供 parsers 使用。
    """
    import io as _io
    from python_calamine import CalamineWorkbook
    from openpyxl import Workbook
    cwb = CalamineWorkbook.from_filelike(_io.BytesIO(content))
    wb = Workbook()
    wb.remove(wb.active)
    for sn in cwb.sheet_names:
        sheet = cwb.get_sheet_by_name(sn)
        # calamine: visible_state 'visible'/'hidden'/'very-hidden' → 仅 visible 保留
        vstate = getattr(sheet, 'visible_state', None) or getattr(sheet, 'visibility', None)
        if vstate and str(vstate).lower() != 'visible':
            continue
        rows = sheet.to_python()
        ws = wb.create_sheet(sn[:31])
        for row in rows:
            ws.append([(v if v != '' else None) for v in row])
    return wb


def _to_list(v):
    if v is None: return []
    if isinstance(v, list): return v
    return json.loads(v)


def _resolve_layer_and_path(cur, source_project_ids, filename):
    """归属判断 → (layer, target_key)"""
    project_ids = list(set(source_project_ids))
    if not project_ids:
        return None, None

    if len(project_ids) == 1:
        # 项目层
        cur.execute("""SELECT p.short_name AS proj_short, e.short_name AS ent_short
                       FROM projects p JOIN enterprises e ON p.enterprise_id=e.id
                       WHERE p.id=%s""", (project_ids[0],))
        row = cur.fetchone()
        if not row:
            return None, None
        proj_short, ent_short = row
        return 'project', cos_ops.resolve_target_path(
            'project', ent_short=ent_short,
            proj_short=proj_short or str(project_ids[0]),
            filename=filename)

    # 多项目：查所有企业 / 实控人
    placeholders = ','.join(['%s'] * len(project_ids))
    cur.execute(f"""SELECT DISTINCT e.id, e.short_name
                    FROM projects p JOIN enterprises e ON p.enterprise_id=e.id
                    WHERE p.id IN ({placeholders})""", project_ids)
    ents = cur.fetchall()
    if len(ents) == 1:
        # 企业层
        return 'enterprise', cos_ops.resolve_target_path(
            'enterprise', ent_short=ents[0][1], filename=filename)

    # 跨企业：查实控人
    ent_ids = [e[0] for e in ents]
    placeholders2 = ','.join(['%s'] * len(ent_ids))
    cur.execute(f"""SELECT DISTINCT c.id, c.name
                    FROM controllers c
                    JOIN controller_enterprise_map m ON m.controller_id=c.id
                    WHERE m.enterprise_id IN ({placeholders2})""", ent_ids)
    ctrls = cur.fetchall()
    if len(ctrls) == 1:
        # 实控人层
        return 'controller', cos_ops.resolve_target_path(
            'controller', controller_name=ctrls[0][1], filename=filename)

    # 跨实控人或无实控人：fallback 项目层第一项
    return 'project_fallback', None


def _load_file_bytes(ai_cos_key, source_urls, expected_hash=None):
    """优先从 ai 桶下载（已归位的）；fallback HTTP GET source_urls[0]。

    expected_hash: raw_files.file_hash。给定时校验 ai 桶内容是否一致——
        不一致说明 ai 桶被旧版本覆盖（sync_cos / relocate bug），自动重新
        从 source_urls 下载并覆盖 ai 桶。
    """
    import hashlib
    body = None
    if ai_cos_key:
        try:
            body = download_bytes(ai_cos_key)
        except Exception:
            body = None
        # 校验 hash：不一致 → 视为 ai 桶污染，强制走 source_urls 重新拉
        if body and expected_hash:
            actual = hashlib.sha256(body).hexdigest()
            if actual != expected_hash:
                print(f'[WARN] ai 桶 hash 不一致 key={ai_cos_key} expected={expected_hash[:16]}.. actual={actual[:16]}.. → fallback source_urls')
                body = None
    if body:
        return body
    if source_urls:
        import requests
        resp = requests.get(source_urls[0], timeout=60)
        resp.raise_for_status()
        body = resp.content
        # source_urls 拉到的视为权威版本：重新覆盖 ai 桶
        if ai_cos_key and (not expected_hash or hashlib.sha256(body).hexdigest() == expected_hash):
            try:
                from scripts._cos import upload_bytes
                upload_bytes(ai_cos_key, body)
                print(f'[fix] 用 source_urls 内容覆盖 ai 桶 key={ai_cos_key}')
            except Exception as e:
                print(f'[WARN] 覆盖 ai 桶失败 key={ai_cos_key}: {e}')
        return body
    raise RuntimeError('no ai_cos_key and no source_urls')


def process_file(cur, fid, override_project_id=None, attendance_collector=None,
                  bill_collector=None, payroll_collector=None, wage_collector=None):
    """处理单个 raw_files 行
    override_project_id: 跨项目挂载的 xlsx 装入指定项目时使用（如 id=50 同时被计时+计件挂载）
    *_collector: 给定时该 kind 的 rows 不立即写库，由 main 在批末统一 flush
        （方案 B 跨文件 dedup + 批量写）
    """
    cur.execute("""SELECT file_hash, ai_cos_key, source_urls, source_filenames,
                          source_project_ids, parse_status
                   FROM raw_files WHERE id=%s""", (fid,))
    row = cur.fetchone()
    if not row:
        return {'status': 'not_found'}
    file_hash, ai_cos_key, urls_j, names_j, projs_j, parse_status = row

    urls = _to_list(urls_j)
    names = _to_list(names_j)
    if override_project_id:
        projs_j = [override_project_id]
    projs = _to_list(projs_j)
    if not projs:
        return {'status': 'no_project_ids'}

    # 1. 下载文件（带 hash 校验，防 ai 桶被旧版本污染）
    try:
        body = _load_file_bytes(ai_cos_key, urls, expected_hash=file_hash)
    except Exception as e:
        cur.execute("""UPDATE raw_files SET parse_status='failed',
                       parse_error=%s, parsed_at=NOW() WHERE id=%s""",
                    (f'download: {type(e).__name__}: {e}'[:500], fid))
        return {'status': 'download_failed', 'error': str(e)}

    # 2. mime 识别
    mime = detect_mime(body)

    # 2b. PDF 分支：按文件名判类型 → 调对应 PDF parser
    if mime == 'pdf':
        filename = (names[0] if names else '').rsplit('/', 1)[-1]
        pdf_kind = _classify_pdf_by_filename(filename)
        if pdf_kind == 'unknown':
            cur.execute("""UPDATE raw_files SET parse_status='skipped',
                           detected_type='pdf', parsed_at=NOW() WHERE id=%s""",
                        (fid,))
            return {'status': 'skip_pdf', 'reason': 'unknown_pdf_kind', 'filename': filename}

        # 取项目配置 + business_cycle + fallback_bm
        project_id = projs[0]
        cur.execute("SELECT enterprise_id FROM projects WHERE id=%s", (project_id,))
        proj = cur.fetchone()
        if not proj:
            cur.execute("""UPDATE raw_files SET parse_status='failed',
                           parse_error='project_not_seeded', parsed_at=NOW() WHERE id=%s""",
                        (fid,))
            return {'status': 'project_missing'}
        enterprise_id = proj[0]
        from etl._utils import get_business_cycle
        business_cycle = get_business_cycle(cur, project_id)

        cur.execute("""SELECT source_bill_ids FROM raw_files WHERE id=%s""", (fid,))
        _row = cur.fetchone()
        bill_ids = _to_list(_row[0] if _row else None)
        fallback_bm = None
        if bill_ids:
            cur.execute("""SELECT bill_month FROM raw_mini_a_bill
                           WHERE id=%s AND bill_month IS NOT NULL LIMIT 1""",
                        (bill_ids[0],))
            b = cur.fetchone()
            if b and b[0]:
                fallback_bm = str(b[0]).split(' ')[0][:7]

        try:
            if pdf_kind == 'attendance':
                r = parse_attendance_pdf.process_pdf(
                    cur, project_id=project_id, enterprise_id=enterprise_id,
                    business_cycle=business_cycle, source_file_id=fid,
                    filename=filename, body=body, fallback_bm=fallback_bm)
            elif pdf_kind == 'bill':
                r = parse_bill_pdf.process_pdf(
                    cur, project_id=project_id, enterprise_id=enterprise_id,
                    business_cycle=business_cycle, source_file_id=fid,
                    filename=filename, body=body, fallback_bm=fallback_bm)
            else:
                cur.execute("""UPDATE raw_files SET parse_status='skipped',
                               detected_type='pdf', parsed_at=NOW() WHERE id=%s""",
                            (fid,))
                return {'status': 'skip_pdf', 'reason': f'parser_for_{pdf_kind}_not_implemented'}

            cur.execute("""UPDATE raw_files SET parse_status='parsed',
                           detected_type=%s, parsed_at=NOW(), parse_error=NULL
                           WHERE id=%s""",
                        (f'pdf_{pdf_kind}', fid))
            return {'status': 'parsed', 'detected_type': f'pdf_{pdf_kind}',
                    'sheets': [{'sheet': 'pdf', 'kind': pdf_kind, **r}]}
        except Exception as e:
            cur.execute("""UPDATE raw_files SET parse_status='failed',
                           parse_error=%s, parsed_at=NOW() WHERE id=%s""",
                        (f'pdf_{pdf_kind}: {type(e).__name__}: {e}'[:500], fid))
            return {'status': 'parser_failed', 'error': str(e)}

    if mime not in ('xlsx_or_zip', 'xls'):
        cur.execute("""UPDATE raw_files SET parse_status='skipped',
                       detected_type=%s, parsed_at=NOW() WHERE id=%s""",
                    (mime, fid))
        return {'status': f'skip_{mime}'}

    # 3. 加载 workbook（xlsx 用 openpyxl，失败 fallback calamine；xls 用 xlrd）
    _filename = (names[0] if names else '').rsplit('/', 1)[-1] if names else None
    try:
        if mime == 'xls':
            wb = _load_xls_as_openpyxl(body, filename=_filename)
        else:
            try:
                # 不用 read_only — read_only 下 iter_rows 在某些 xlsx 上行为异常（顺丰转账明细只返回 col 0）
                wb = load_workbook(io.BytesIO(body), data_only=True)
            except Exception as e_op:
                # openpyxl 解析失败（如样式表损坏）→ fallback calamine
                wb = _load_xlsx_via_calamine(body)
    except Exception as e:
        cur.execute("""UPDATE raw_files SET parse_status='failed',
                       parse_error=%s, parsed_at=NOW() WHERE id=%s""",
                    (f'load_{mime}: {type(e).__name__}: {e}'[:500], fid))
        return {'status': 'parse_failed', 'error': str(e)}

    # 4. 识别（v2: 项目级 DB 规则；缺规则自动 fallback DEFAULT_RULES，与旧引擎等价）
    _proj_for_id = projs[0] if projs else None
    ident = identify(wb, project_id=_proj_for_id, conn=cur.connection)

    # 4.5 unknown sheet 入 pending_classify_sheets 待用户在 UI 上加规则
    if _proj_for_id:
        for sh in ident['sheets']:
            if sh['kind'] != 'unknown':
                continue
            try:
                ws = wb[sh['name']]
                preview = []
                for row in ws.iter_rows(max_row=4, values_only=True):
                    preview.append(' | '.join(str(v).strip() if v is not None else '' for v in row))
                cur.execute("""INSERT IGNORE INTO pending_classify_sheets
                                 (project_id, raw_file_id, sheet_name, headers_preview)
                               VALUES (%s, %s, %s, %s)""",
                            (_proj_for_id, fid, sh['name'][:255], '\n'.join(preview)[:4000]))
            except Exception:
                pass  # pending 写失败不影响主流程

    # 5. 归属判断 + COS 归位（仅"考账发工"类需要归位；unknown 留 _inbox/）
    new_cos_key = ai_cos_key
    if ident['kao_zhang_fa_gong']:
        filename = (names[0] if names else f'{file_hash}.xlsx')
        # 文件名前去掉 _inbox/ 前缀的时间戳？这里保留原文件名
        if '/' in filename:
            filename = filename.rsplit('/', 1)[-1]
        layer, target_key = _resolve_layer_and_path(cur, projs, filename)
        if target_key:
            try:
                # 防止同名不同 hash 文件归位时 COS 被覆盖丢失:
                # 目标 key 已存在 → 看是否被本 hash 占用,不是则加 hash 后缀
                if cos_ops.exists(target_key):
                    cur.execute("SELECT id, file_hash FROM raw_files WHERE ai_cos_key=%s",
                                (target_key,))
                    existing = cur.fetchone()
                    if existing and existing[1] != file_hash:
                        # 同名不同内容 → 加 hash 前 8 位后缀
                        suffix = file_hash[:8]
                        if '.' in target_key:
                            base, ext = target_key.rsplit('.', 1)
                            target_key = f'{base}_{suffix}.{ext}'
                        else:
                            target_key = f'{target_key}_{suffix}'
                action = cos_ops.relocate(ai_cos_key, target_key)
                cos_ops.update_raw_files_key(cur, fid, target_key)
                new_cos_key = target_key
            except Exception as e:
                # 归位失败不阻塞 mart 写入
                cur.execute("""UPDATE raw_files SET parse_error=%s WHERE id=%s""",
                            (f'cos_relocate_failed: {e}'[:500], fid))

    # 6. 项目配置（按首个 project_id）
    project_id = projs[0]
    cur.execute("SELECT enterprise_id FROM projects WHERE id=%s", (project_id,))
    proj = cur.fetchone()
    if not proj:
        cur.execute("""UPDATE raw_files SET parse_status='failed',
                       parse_error='project_not_seeded', parsed_at=NOW() WHERE id=%s""",
                    (fid,))
        return {'status': 'project_missing'}
    enterprise_id = proj[0]
    from etl._utils import get_business_cycle
    business_cycle = get_business_cycle(cur, project_id)

    # 6b. fallback_bm：从挂载的 mini_a_bill.bill_month 取（首个）
    cur.execute("""SELECT id, source_bill_ids FROM raw_files WHERE id=%s""", (fid,))
    _row = cur.fetchone()
    bill_ids = _to_list(_row[1] if _row else None)
    fallback_bm = None
    if bill_ids:
        cur.execute("""SELECT bill_month FROM raw_mini_a_bill
                       WHERE id=%s AND bill_month IS NOT NULL LIMIT 1""",
                    (bill_ids[0],))
        b = cur.fetchone()
        if b and b[0]:
            # bill_month 可能是 'YYYY-MM' 或 'YYYY-MM ~ YYYY-MM'
            fallback_bm = str(b[0]).split(' ')[0][:7]

    # 7. 按 sheet kind dispatch（支持同 sheet 多 kind 装入：matches 列表）
    # payroll 类:对 source_project_ids 里每个 project 循环跑一次（各自按规则装入）
    # 其他类(attendance/bill/wage_sheet):只跑首个 project_id（同企业归位逻辑已处理跨项目）
    sheet_results = []
    failed = False
    for s in ident['sheets']:
        sname = s['name']
        # matches 由 identifier 层透传；为空（empty/unknown）则跳过
        matches = s.get('matches') or []

        # === 主 loop：每个命中的 kind 各装一遍（matches 已按 KIND_ORDER 排好序）===
        for match in matches:
            kind = match.get('kind')
            if kind not in KIND_TO_PARSER:
                sheet_results.append({'sheet': sname, 'kind': kind, 'skipped': 'no_parser'})
                continue
            parser = KIND_TO_PARSER[kind]
            ws = wb[sname]

            # 决定本 sheet 要跑的 project_id 列表
            if kind == 'payroll' and len(projs) > 1:
                target_pids = list(projs)
            else:
                target_pids = [project_id]

            for pid in target_pids:
                cur.execute("SELECT enterprise_id FROM projects WHERE id=%s", (pid,))
                _r = cur.fetchone()
                if not _r:
                    sheet_results.append({'sheet': sname, 'kind': kind,
                                           'project_id': pid, 'skipped': 'project_not_seeded'})
                    continue
                ent_id_p = _r[0]
                cycle_p = get_business_cycle(cur, pid)

                kwargs = dict(project_id=pid, enterprise_id=ent_id_p,
                              business_cycle=cycle_p,
                              source_file_id=fid, sheet_name=sname, ws=ws,
                              format_id=match.get('format_id'))
                if kind in ('bill', 'wage_sheet', 'attendance'):
                    kwargs['fallback_bm'] = fallback_bm

                # === by-handler 路由：standard / two_region_attendance / pivot_attendance / aggregate_label ===
                handler_name = match.get('handler')
                cm = match.get('column_mapping')

                # format.handler 优先覆盖 classify rule.handler
                # （format 管理页一处改即生效，无需逐条 classify rule 改 handler）
                if match.get('format_id'):
                    try:
                        cur.execute('SELECT handler FROM project_formats WHERE id=%s',
                                    (int(match['format_id']),))
                        _r = cur.fetchone()
                        if _r and _r[0]:
                            handler_name = _r[0]
                    except Exception:
                        pass

                # 聚合标签账单：bill kind 只要项目配了 aggregate rule 就走聚合（无需 handler='aggregate_label'）
                # 聚合 logic 仅覆盖 bill_totals.amount；bill_persons 仍按标准 person 解析照常入库
                if kind == 'bill':
                    try:
                        from etl.views.aggregate_label_rules import load_active_rules
                        _agg_rules = load_active_rules(
                            cur, pid, format_id=match.get('format_id'))
                        if _agg_rules:
                            kwargs['aggregate_rules'] = _agg_rules
                    except Exception as e_a:
                        sheet_results.append({'sheet': sname, 'kind': kind, 'project_id': pid,
                                               'note': f'aggregate_route_fallback: {e_a}'})

                if (USE_HANDLER_ROUTE
                        and kind in ('attendance', 'bill', 'wage_sheet', 'payroll')
                        and handler_name in ('standard', 'two_region_attendance', 'pivot_attendance')
                        and (cm or handler_name == 'pivot_attendance')):
                    try:
                        if handler_name == 'two_region_attendance':
                            if kind != 'attendance':
                                raise ValueError(f"two_region_attendance handler 仅支持 kind='attendance'")
                            from etl.parsers.handlers.two_region_attendance import parse as tra_parse
                            std_rows = tra_parse(ws, column_mapping=cm,
                                                  sheet_name=sname,
                                                  business_month=fallback_bm,
                                                  fallback_bm=fallback_bm)
                        elif handler_name == 'pivot_attendance':
                            if kind != 'attendance':
                                raise ValueError(f"pivot_attendance handler 仅支持 kind='attendance'")
                            from etl.views.pivot_template import get_pivot_template
                            from etl.parsers.handlers.pivot_attendance import parse as piv_parse
                            _config = get_pivot_template(pid, 'attendance',
                                                          format_id=match.get('format_id'))
                            # 透传 business_cycle + business_month，让 pivot 自动推业务周期起点
                            # （sheet 无年月文本时 fallback 用）
                            _config['business_cycle'] = cycle_p
                            _config['business_month'] = fallback_bm
                            std_rows = piv_parse(ws, config=_config, sheet_name=sname)
                        else:
                            from etl.parsers.handlers.standard import parse as std_parse
                            from etl.parsers.handlers._extra_data_augment import collect_extra_columns
                            extra_ctx = {}
                            if kind == 'payroll' and cm.get('pay_time') == '$bill_month':
                                from etl.parsers.parse_payroll_xlsx import _lookup_bill_month_for_file
                                extra_ctx['bill_month'] = _lookup_bill_month_for_file(cur, fid)
                            # 自动透传所有过滤/校验/归属规则用到的列名 → row.extra_data
                            auto_extra = collect_extra_columns(
                                cur, pid, kind, format_id=match.get('format_id'))
                            std_rows = std_parse(ws, kind=kind, column_mapping=cm,
                                                  sheet_name=sname,
                                                  business_month=fallback_bm,
                                                  fallback_bm=fallback_bm,
                                                  auto_extra_columns=auto_extra,
                                                  **extra_ctx)
                        if kind == 'attendance':
                            from etl.parsers.handlers.enterprise_filter import apply_enterprise_filter
                            std_rows, _drop = apply_enterprise_filter(
                                std_rows, kind=kind, project_id=pid, conn=cur.connection,
                                column_mapping=cm or {},
                                format_id=match.get('format_id'))
                        # 所有 kind 跑项目归属（project scope）列过滤
                        # category 映射：attendance/bill → kaoqin_bill；wage_sheet → wage；payroll → payroll
                        from etl._attribution import apply_row_filter_to_dicts
                        _cat = {'attendance':'kaoqin_bill','bill':'kaoqin_bill',
                                'wage_sheet':'wage','payroll':'payroll'}.get(kind)
                        if _cat:
                            std_rows, _drop_attr = apply_row_filter_to_dicts(
                                std_rows, project_id=pid, category=_cat, scope='project',
                                conn=cur.connection, format_id=match.get('format_id'))
                        kwargs['precomputed_rows'] = std_rows
                    except Exception as e_h:
                        sheet_results.append({'sheet': sname, 'kind': kind, 'project_id': pid,
                                               'note': f'handler_route_fallback: {e_h}'})

                # collector 模式（方案 B 跨文件 dedup）：按 kind 注入对应 collector
                _collector_for_kind = {
                    'attendance': attendance_collector,
                    'bill': bill_collector,
                    'payroll': payroll_collector,
                    'wage_sheet': wage_collector,
                }.get(kind)
                if _collector_for_kind is not None:
                    kwargs['collector'] = _collector_for_kind
                try:
                    r = parser.process_sheet(cur, **kwargs)
                    sheet_results.append({'sheet': sname, 'kind': kind,
                                           'project_id': pid,
                                           'handler': handler_name, **r})
                except Exception as e:
                    failed = True
                    sheet_results.append({'sheet': sname, 'kind': kind,
                                           'project_id': pid,
                                           'error': f'{type(e).__name__}: {e}'})
                    break

            if failed:
                break  # 跳出 matches loop

        if failed:
            break  # 跳出 sheets loop

    if failed:
        return {'status': 'parser_failed', 'sheets': sheet_results}

    # 8. 标 parsed
    cur.execute("""UPDATE raw_files SET parse_status='parsed',
                   detected_type=%s, parsed_at=NOW(), parse_error=NULL
                   WHERE id=%s""",
                (ident['file_detected_type'], fid))
    return {'status': 'parsed',
            'detected_type': ident['file_detected_type'],
            'cos_key': new_cos_key,
            'sheets': sheet_results}


def _dedup_by_bill_and_type(cur, fids):
    """同一 (bill_id, detected_type, **basename**) 桶里只保留 first_uploaded_at 最新的。
    basename = source_filenames[0] 去 zip 路径后的最后一段文件名（保留时间戳后缀）。
    场景：用户同 bill 多次上传同名"4月.xlsx" → 只留最新。不同名/不同时间戳导出 → 都保留。
    detected_type=NULL 或 source_bill_ids=空 的不参与去重（保留全部，下次拉再去重）。
    返回 (keep_fids, superseded_fids) — 后者是被新版本取代、本次跳过解析的旧版本。"""
    if not fids:
        return [], []
    placeholders = ','.join(['%s'] * len(fids))
    cur.execute(f"""SELECT id, source_bill_ids, detected_type, first_uploaded_at, source_filenames
                    FROM raw_files WHERE id IN ({placeholders})""", fids)
    rows = cur.fetchall()
    buckets = {}
    free = set()
    for fid, bills_j, dt, t, names_j in rows:
        try:
            bills = json.loads(bills_j) if isinstance(bills_j, str) else (bills_j or [])
        except Exception:
            bills = []
        try:
            names = json.loads(names_j) if isinstance(names_j, str) else (names_j or [])
        except Exception:
            names = []
        if not dt or not bills:
            free.add(fid)
            continue
        # basename = 文件名最后一段（去 zip 路径前缀）
        first_name = names[0] if names else ''
        basename = first_name.replace('\\', '/').rsplit('/', 1)[-1]
        for b in bills:
            buckets.setdefault((str(b), dt, basename), []).append((t, fid))
    superseded = set()
    in_bucket = set()
    for items in buckets.values():
        items.sort(key=lambda x: (x[0] or 0), reverse=True)
        for i, (_, f) in enumerate(items):
            in_bucket.add(f)
            if i > 0 and f != items[0][1]:
                superseded.add(f)
    # 一个文件可能在 A 桶被超越、在 B 桶最新 → 取并集口径"任一桶最新即保留"
    actually_superseded = superseded - {items[0][1] for items in buckets.values()}
    keep = [f for f in fids if f in free or f not in actually_superseded]
    return keep, sorted(actually_superseded)


def _purge_mart_by_source_files(cur, fids):
    """删除 mart 表中 source_file_id 属于 fids 的行（清理旧版本装入痕迹）"""
    if not fids:
        return {}
    placeholders = ','.join(['%s'] * len(fids))
    out = {}
    for tb in ('attendance', 'attendance_summary', 'bill_totals', 'bill_persons',
                'payrolls', 'wage_sheets'):
        try:
            cur.execute(f"DELETE FROM {tb} WHERE source_file_id IN ({placeholders})", fids)
            out[tb] = cur.rowcount
        except Exception as e:
            out[tb] = f'err:{e}'
    return out


def main(force=False, file_id=None, project_id=None, limit=None, override_project_id=None,
         business_month=None):
    """business_month: 'YYYY-MM' 时只解析 source_bill_ids 关联到该业务月 raw_mini_a_bill 的 raw_files"""
    conn = connect('fish-test')
    cur = conn.cursor()

    sql = "SELECT id FROM raw_files WHERE 1=1"
    args = []
    if file_id:
        sql += " AND id=%s"; args.append(file_id)
    elif force:
        sql += " AND parse_status IN ('pending','parsed','failed','skipped','extracted')"
    else:
        sql += " AND parse_status IN ('pending','extracted')"
    if project_id:
        sql += " AND JSON_CONTAINS(source_project_ids, %s)"
        args.append(json.dumps(project_id))
    if business_month:
        # 只跑 source_bill_ids 里有任一 raw_mini_a_bill.bill_month 命中业务月的文件
        sql += """ AND EXISTS (
                     SELECT 1 FROM raw_mini_a_bill b
                     WHERE JSON_CONTAINS(raw_files.source_bill_ids, CAST(b.id AS JSON))
                       AND (b.bill_month=%s OR b.bill_month LIKE %s)
                   )"""
        args += [business_month, f'{business_month}%']
    sql += " ORDER BY first_uploaded_at ASC"
    if limit:
        sql += f" LIMIT {int(limit)}"

    cur.execute(sql, args)
    fids = [r[0] for r in cur.fetchall()]
    print(f'初选 raw_files: {len(fids)} 个')

    # 去重：同 bill_id + detected_type 只跑最新的
    if not file_id:  # 单文件强制模式不去重
        fids, superseded = _dedup_by_bill_and_type(cur, fids)
        if superseded:
            print(f'按 (bill_id+detected_type) 去重，跳过 {len(superseded)} 个旧版本: {superseded}')
            purge_n = _purge_mart_by_source_files(cur, superseded)
            conn.commit()
            print(f'  清理旧版本 mart 行: {purge_n}')
            # 旧版本 mark 'superseded'
            placeholders = ','.join(['%s'] * len(superseded))
            cur.execute(f"""UPDATE raw_files SET parse_status='superseded',
                            parse_error='superseded by newer file in same bill',
                            parsed_at=NOW() WHERE id IN ({placeholders})""", superseded)
            conn.commit()
    print(f'实际处理 raw_files: {len(fids)} 个')

    # project_id 模式（pull_data 单项目拉取）：自动 override，让 process_file 只跑该项目
    # 不要循环 source_project_ids 把别的项目也跑了
    effective_override = override_project_id or project_id

    # 方案 B：4 kind 跨文件 collector，统一 dedup + 批量写
    from etl.loaders.attendance_loader import AttendanceCollector
    from etl.loaders.bill_loader import BillCollector
    from etl.loaders.payroll_loader import PayrollCollector
    from etl.loaders.wage_loader import WageCollector
    att_collector = AttendanceCollector()
    bill_collector = BillCollector()
    payroll_collector = PayrollCollector()
    wage_collector = WageCollector()

    n_ok = n_fail = n_skip = 0
    for fid in fids:
        try:
            result = process_file(cur, fid, override_project_id=effective_override,
                                   attendance_collector=att_collector,
                                   bill_collector=bill_collector,
                                   payroll_collector=payroll_collector,
                                   wage_collector=wage_collector)
            conn.commit()
            status = result.get('status', 'unknown')
            print(f'  id={fid:4d} [{status:18s}] {result.get("detected_type", "")}')
            for sh in result.get('sheets', []):
                print(f'        {sh}')
            if status == 'parsed': n_ok += 1
            elif status.startswith('skip'): n_skip += 1
            else: n_fail += 1
        except Exception as e:
            conn.rollback()
            n_fail += 1
            print(f'  id={fid:4d} [EXCEPTION] {type(e).__name__}: {e}')
            try:
                cur.execute("""UPDATE raw_files SET parse_status='failed',
                               parse_error=%s, parsed_at=NOW() WHERE id=%s""",
                            (f'{type(e).__name__}: {e}'[:500], fid))
                conn.commit()
            except Exception:
                conn.rollback()

    # 批末 flush：4 kind 跨文件 dedup + 批量 INSERT
    for label, col in (('attendance', att_collector), ('bill', bill_collector),
                       ('payroll', payroll_collector), ('wage_sheet', wage_collector)):
        if col.entries:
            flush_result = col.flush(cur)
            conn.commit()
            print(f'\n[{label} batch flush] {flush_result}')

    print(f'\n汇总: OK={n_ok}, SKIP={n_skip}, FAIL={n_fail}')


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--force', action='store_true')
    ap.add_argument('--file-id', type=int)
    ap.add_argument('--project-id', type=int)
    ap.add_argument('--limit', type=int)
    ap.add_argument('--override-project-id', type=int,
                    help='强制把所有处理的 xlsx 装入指定项目（跨项目挂载场景用）')
    ap.add_argument('--business-month',
                    help='YYYY-MM；只解析关联到该业务月 raw_mini_a_bill 的 raw_files')
    args = ap.parse_args()
    main(force=args.force, file_id=args.file_id, project_id=args.project_id, limit=args.limit,
         override_project_id=args.override_project_id,
         business_month=args.business_month)
