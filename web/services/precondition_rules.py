"""通用前置规则1-7"""
import os
import re
import openpyxl
from datetime import date, datetime, timedelta
from config import BASE_DATA_DIR


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    # 处理 "2024-04-01至2026-03-31" 格式，取结束日期
    if '至' in s:
        s = s.split('至')[-1].strip()
    if s.isdigit() and len(s) == 8:
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except Exception:
            return None
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日']:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def get_header_map(ws):
    result = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            result[str(val).strip()] = c
    return result


def contains(haystack, needle):
    if not haystack or not needle:
        return False
    h, n = str(haystack).strip(), str(needle).strip()
    return n in h or h in n


def load_controller_info(company):
    """加载实控人信息和授信额度"""
    filepath = os.path.join(BASE_DATA_DIR, "实控人关系.xlsx")
    if not os.path.isfile(filepath):
        return [], {}, 0

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 找实控人
    ws = wb["实控人-公司"]
    controllers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ctrl, comp = row[0], row[1]
        if comp and contains(comp, company) and ctrl not in controllers:
            controllers.append(ctrl)

    # 找名下所有公司
    all_companies = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ctrl, comp = row[0], row[1]
        if ctrl in controllers:
            all_companies.setdefault(ctrl, []).append(comp)

    # 授信额度(万元)
    ws_auth = wb["授信"]
    credit_limit = 0
    for row in ws_auth.iter_rows(min_row=2, values_only=True):
        if row[0] in controllers and row[1]:
            credit_limit = float(row[1])

    # 临时占用
    temp_occupy = 0
    if "临时占用" in wb.sheetnames:
        ws_temp = wb["临时占用"]
        for row in ws_temp.iter_rows(min_row=2, values_only=True):
            if row[0] in controllers and row[4] and str(row[6]).strip() != "已释放":
                try:
                    temp_occupy += float(row[4])
                except Exception:
                    pass

    wb.close()
    return controllers, all_companies, credit_limit, temp_occupy


def load_project_info_full(company, project):
    """读取项目信息.xlsx"""
    info_file = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "项目信息.xlsx")
    if not os.path.isfile(info_file):
        return {}
    try:
        wb = openpyxl.load_workbook(info_file, data_only=True)
        ws = wb.worksheets[0]
        info = {}
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                info[str(row[0]).strip()] = row[1]
        wb.close()
        return info
    except Exception:
        return {}


def load_return_days(company, project):
    """从出款管理表读取回款天数"""
    mgr_file = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "出款管理", "出款管理表.xlsx")
    if not os.path.isfile(mgr_file):
        return None
    try:
        wb = openpyxl.load_workbook(mgr_file, data_only=True)
        if "回款时间配置" not in wb.sheetnames:
            wb.close()
            return None
        ws = wb["回款时间配置"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                wb.close()
                return int(row[2])
        wb.close()
    except Exception:
        pass
    return None


def load_total_in_loan(controllers, all_companies):
    """汇总实控人名下所有在贷金额"""
    total = 0
    for ctrl in controllers:
        for comp in all_companies.get(ctrl, []):
            comp_dir = os.path.join(BASE_DATA_DIR, "企业", comp)
            if not os.path.isdir(comp_dir):
                continue
            proj_dir = os.path.join(comp_dir, "项目")
            if not os.path.isdir(proj_dir):
                continue
            for proj in os.listdir(proj_dir):
                mgr_file = os.path.join(proj_dir, proj, "出款管理", "出款管理表.xlsx")
                if not os.path.isfile(mgr_file):
                    continue
                try:
                    wb = openpyxl.load_workbook(mgr_file, data_only=True)
                    if "出款记录" not in wb.sheetnames:
                        wb.close()
                        continue
                    ws = wb["出款记录"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        amount = row[3]  # 出款金额
                        status = str(row[8]) if row[8] else ""  # 回款状态
                        if amount and "已回款" not in status:
                            try:
                                total += float(amount)
                            except Exception:
                                pass
                    wb.close()
                except Exception:
                    continue
    return total


def load_enterprise_credit(company):
    """加载企业征信"""
    credit_dir = os.path.join(BASE_DATA_DIR, "企业", company, "征信")
    if not os.path.isdir(credit_dir):
        return None, []

    query_time = None
    loans = []

    for fname in os.listdir(credit_dir):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        fpath = os.path.join(credit_dir, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for sn in wb.sheetnames:
                if sn == "模版":
                    continue
                ws = wb[sn]
                hmap = get_header_map(ws)
                if ws.max_row >= 2:
                    qt = parse_date(ws.cell(row=2, column=1).value)
                    if qt and (query_time is None or qt < query_time):
                        query_time = qt

                gl_col = hmap.get("关联负债")
                if not gl_col:
                    continue
                dq_col = hmap.get("到期日期")
                bal_col = hmap.get("贷款余额")
                bank_col = hmap.get("贷款银行", 2)
                jq_col = hmap.get("是否结清")

                for r in range(2, ws.max_row + 1):
                    if jq_col and str(ws.cell(row=r, column=jq_col).value) == "是":
                        continue
                    if str(ws.cell(row=r, column=gl_col).value) == "是":
                        loans.append({
                            "来源": fname.replace('.xlsx', ''),
                            "类型": "企业",
                            "银行": str(ws.cell(row=r, column=bank_col).value or "")[:30],
                            "余额": ws.cell(row=r, column=bal_col).value if bal_col else None,
                            "到期日期": parse_date(ws.cell(row=r, column=dq_col).value) if dq_col else None,
                        })
            wb.close()
        except Exception:
            continue

    return query_time, loans


def load_personal_credit(controllers):
    """加载个人征信"""
    query_times = {}
    loans = []

    for ctrl in controllers:
        credit_dir = os.path.join(BASE_DATA_DIR, "个人", ctrl, "征信")
        if not os.path.isdir(credit_dir):
            continue
        for fname in os.listdir(credit_dir):
            if not fname.endswith('.xlsx') or fname.startswith('~$'):
                continue
            fpath = os.path.join(credit_dir, fname)
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb.worksheets[0]
                hmap = get_header_map(ws)
                if ws.max_row >= 2:
                    qt = parse_date(ws.cell(row=2, column=1).value)
                    if qt:
                        query_times[ctrl] = qt

                gl_col = hmap.get("关联负债")
                if gl_col:
                    dq_col = hmap.get("到期日期")
                    bal_col = hmap.get("贷款余额")
                    bank_col = hmap.get("贷款机构", 2)
                    jq_col = hmap.get("是否结清")

                    for r in range(2, ws.max_row + 1):
                        if jq_col and str(ws.cell(row=r, column=jq_col).value) == "是":
                            continue
                        if str(ws.cell(row=r, column=gl_col).value) == "是":
                            loans.append({
                                "来源": ctrl,
                                "类型": "个人",
                                "银行": str(ws.cell(row=r, column=bank_col).value or "")[:30],
                                "余额": ws.cell(row=r, column=bal_col).value if bal_col else None,
                                "到期日期": parse_date(ws.cell(row=r, column=dq_col).value) if dq_col else None,
                            })
                wb.close()
            except Exception:
                continue

    return query_times, loans


def run_precondition_checks(proj_info, year, month, today, controllers, all_companies, company):
    """执行前置规则1-7，返回 (checks, blocked, warnings)"""
    checks = []
    blocked = []
    warnings = []

    service_start = date(year, month, 1)

    # 回款时间
    return_days = load_return_days(company, proj_info.get('_project', ''))
    return_date = service_start + timedelta(days=return_days) if return_days else None

    # 规则1: 甲方合同有效期
    contract_end = parse_date(proj_info.get("甲方合同有效期"))
    if contract_end is None:
        checks.append({"规则": "规则1-甲方合同有效期", "内容": "未填写", "结果": "风险提示"})
    elif contract_end < service_start:
        checks.append({"规则": "规则1-甲方合同有效期", "内容": f"{contract_end}，已过期", "结果": "不通过"})
        blocked.append(f"甲方合同已过期（{contract_end}）")
    elif contract_end.year == year and contract_end.month == month:
        checks.append({"规则": "规则1-甲方合同有效期", "内容": f"{contract_end}，账期在最后一个月", "结果": "不通过"})
        blocked.append(f"账期{month}月在甲方合同最后月（到期{contract_end}）")
    else:
        checks.append({"规则": "规则1-甲方合同有效期", "内容": f"{contract_end}", "结果": "通过"})

    # 规则2: 回款时间
    if return_date is None:
        checks.append({"规则": "规则2-回款时间", "内容": "未配置", "结果": "风险提示"})
    elif return_date < today:
        checks.append({"规则": "规则2-回款时间", "内容": f"{return_date}早于今天", "结果": "不通过"})
        blocked.append(f"回款时间{return_date}早于今天")
    else:
        checks.append({"规则": "规则2-回款时间", "内容": str(return_date), "结果": "通过"})

    # 规则3: 保理合同有效期
    factoring_end = parse_date(proj_info.get("保理合同有效期"))
    if factoring_end is None:
        checks.append({"规则": "规则3-保理合同有效期", "内容": "未填写", "结果": "风险提示"})
    else:
        diff_days = (factoring_end - today).days
        if diff_days <= 0:
            checks.append({"规则": "规则3-保理合同有效期", "内容": f"已过期（{factoring_end}）", "结果": "不通过"})
            blocked.append(f"保理合同已过期（{factoring_end}）")
        elif diff_days <= 30:
            checks.append({"规则": "规则3-保理合同有效期", "内容": f"距今{diff_days}天（<1月）", "结果": "不通过"})
            blocked.append(f"保理合同距到期仅{diff_days}天")
        else:
            checks.append({"规则": "规则3-保理合同有效期", "内容": f"距今{diff_days}天", "结果": "通过"})

    # 规则4: 企业征信时效
    ent_query_time, ent_loans = load_enterprise_credit(company)
    if ent_query_time is None:
        checks.append({"规则": "规则4-企业征信时效", "内容": "未找到企业征信", "结果": "风险提示"})
    else:
        diff_days = (today - ent_query_time).days
        if diff_days > 180:
            checks.append({"规则": "规则4-企业征信时效", "内容": f"{ent_query_time}距今{diff_days}天", "结果": "不通过"})
            blocked.append(f"企业征信距今{diff_days}天，超过6个月")
        else:
            checks.append({"规则": "规则4-企业征信时效", "内容": f"{ent_query_time}距今{diff_days}天", "结果": "通过"})

    # 规则5: 个人征信时效
    per_query_times, per_loans = load_personal_credit(controllers)
    if not per_query_times:
        checks.append({"规则": "规则5-个人征信时效", "内容": "未找到个人征信", "结果": "风险提示"})
    else:
        for ctrl, qt in per_query_times.items():
            diff_days = (today - qt).days
            if diff_days > 180:
                checks.append({"规则": f"规则5-个人征信时效（{ctrl}）", "内容": f"{qt}距今{diff_days}天", "结果": "不通过"})
                blocked.append(f"{ctrl}个人征信距今{diff_days}天，超过6个月")
            else:
                checks.append({"规则": f"规则5-个人征信时效（{ctrl}）", "内容": f"{qt}距今{diff_days}天", "结果": "通过"})

    # 规则6: 贷款到期评估
    all_loans = ent_loans + per_loans
    if all_loans and return_date:
        for loan in all_loans:
            due = loan["到期日期"]
            bank = loan["银行"]
            if due and return_date > due:
                gap = (return_date - due).days
                blocked.append(f"{bank}贷款到期{due}，早于回款{return_date}（{gap}天）")
                checks.append({"规则": "规则6-贷款到期", "内容": f"{bank}到期{due}，回款晚{gap}天", "结果": "不通过"})
            elif due and return_date:
                gap = (due - return_date).days
                checks.append({"规则": "规则6-贷款到期", "内容": f"{bank}距到期{gap}天", "结果": "通过"})
    elif not all_loans:
        checks.append({"规则": "规则6-贷款到期评估", "内容": "无关联负债", "结果": "跳过"})

    # 规则7: 保险合规
    insurance = proj_info.get("保险合规")
    if insurance and str(insurance).strip() == "是":
        checks.append({"规则": "规则7-保险合规", "内容": "是", "结果": "通过"})
    else:
        checks.append({"规则": "规则7-保险合规", "内容": str(insurance), "结果": "风险提示"})

    return checks, blocked, warnings, return_date, return_days
