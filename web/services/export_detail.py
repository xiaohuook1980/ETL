"""出款计算明细导出：6 sheet xlsx
sheets: 考勤预估 / 账单金额 / 发薪流水 / 工资表 / 考勤×发薪 / 考勤×工资表
"""
import io
import sys
from datetime import date as _date, datetime
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scripts._db import connect


HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')
HEADER_FONT = Font(bold=True)
TOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')
TOTAL_FONT = Font(bold=True)


def _ws_init(wb, name, headers, widths=None):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    return ws


def _append_total(ws, label_col, label, sum_cols):
    """给 ws 末尾加合计行。sum_cols: list of (col_idx, value)"""
    r = ws.max_row + 1
    cell = ws.cell(row=r, column=label_col, value=label)
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    for ci, v in sum_cols:
        c = ws.cell(row=r, column=ci, value=v)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        if isinstance(v, (int, float)):
            c.number_format = '#,##0.00'


def _set_money_col(ws, col_idx):
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=col_idx)
        if isinstance(c.value, (int, float)):
            c.number_format = '#,##0.00'


def _file_label(cur, file_id):
    if not file_id:
        return ''
    cur.execute("SELECT source_filenames FROM raw_files WHERE id=%s", (file_id,))
    r = cur.fetchone()
    if not r or not r[0]:
        return f'fid={file_id}'
    import json
    try:
        names = json.loads(r[0]) if isinstance(r[0], str) else r[0]
        return (names[0] if names else '').rsplit('/', 1)[-1]
    except Exception:
        return f'fid={file_id}'


NORMAL_ROWS_LABELS = [
    (1,  '#1_甲方账单',         '甲方账单金额'),
    (2,  '#2_账单出款上限',     '账单:出款上限'),
    (3,  '#3_发薪流水金额',     '发薪流水:金额'),
    (4,  '#4_已直发',           '发薪流水:已直发'),
    (5,  '#5_发薪出款上限',     '发薪流水:出款上限'),
    (6,  '#6_工资表结算工资',   '工资表:结算工资'),
    (7,  '#7_本月匹配率',       '工资表:本月匹配率'),
    (8,  '#8_工资表匹配金额',   '工资表:匹配金额'),
    (9,  '#9_工资表预计直发',   '工资表:预计直发'),
    (10, '#10_上月匹配率',      '工资表:上月匹配率'),
    (11, '#11_工资表出款上限',  '工资表:出款上限'),
    (12, '#12_本周期已垫付',    '本业务周期已垫付'),
    (13, '#13_本项目出款金额',  '本项目出款金额'),
    (14, '#14_代收超额扣减',    '代收超额扣减'),
    (15, '#15_授信余额',        '授信余额'),
    (16, '#16_客户申请金额',    '客户申请金额'),
    (17, '#17_最终出款',        '最终出款'),
]
PREPAY_ROWS_LABELS = [
    (1,  '#1_考勤账单金额',    '考勤账单金额'),
    (2,  '#2_预付考勤估计',    '预付考勤估计'),
    (3,  '#3_合计',            '合计'),
    (4,  '#4_考勤出款上限',    '考勤出款上限'),
    (5,  '#5_已发薪金额',      '已发薪金额'),
    (6,  '#6_已直发',          '已直发(验证)'),
    (7,  '#7_已垫付',          '已垫付'),
    (8,  '#8_账户结余',        '账户结余'),
    (9,  '#9_项目代收超额',    '项目代收超额'),
    (10, '#10_出款金额',       '出款金额'),
    (11, '#11_代收超额扣减',   '代收超额扣减(实控人)'),
    (12, '#12_授信余额',       '授信余额'),
    (13, '#13_最终出款',       '最终出款'),
]


def export_detail(project_id, business_month, apply_date=None):
    """返回 BytesIO 文件流"""
    project_id = int(project_id)
    conn = connect('fish-test')
    cur = conn.cursor()

    cur.execute("""SELECT p.id, p.title, e.short_name, p.daily_deduction_hours,
                          p.profit_ratio, p.finance_mode
                   FROM projects p JOIN enterprises e ON e.id=p.enterprise_id
                   WHERE p.id=%s""", (project_id,))
    meta = cur.fetchone()
    proj_title = meta[1] if meta else ''
    ent_short = meta[2] if meta else ''
    daily_ded = float(meta[3] or 0) if meta else 0
    profit = float(meta[4] or 0.8) if meta else 0.8
    finance_mode = (meta[5] or 'normal') if meta else 'normal'

    cur.execute("""SELECT shift_name, price, unit FROM unit_prices
                   WHERE project_id=%s ORDER BY (shift_name<>''), id""", (project_id,))
    price_map = {}
    default_price = (0.0, '元/小时')
    has_default = False
    for sn, p, u in cur.fetchall():
        if sn:
            price_map[sn] = (float(p or 0), u or '元/小时')
        elif not has_default:
            default_price = (float(p or 0), u or '元/小时')
            has_default = True

    def get_price(shift_name):
        if shift_name and shift_name in price_map:
            return price_map[shift_name]
        return default_price

    wb = Workbook()
    wb.remove(wb.active)

    # ============ Sheet 0: 出款结果（最前） ============
    try:
        from services.calc_adapter import run_calc
        mode = 'prepay' if finance_mode == 'prepay' else 'normal'
        calc_out = run_calc(project_id, business_month,
                            apply_date=apply_date, mode=mode)
        detail = calc_out.get('detail', {}) if calc_out else {}
        calc_err = calc_out.get('error') if calc_out else 'no_output'
    except Exception as e:
        detail = {}
        calc_err = f'{type(e).__name__}: {e}'

    ws_r = _ws_init(wb, '出款结果',
                     ['#', '输出项', '金额', '说明'],
                     [6, 22, 16, 80])
    # 顶部基本信息
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='项目').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=f'{ent_short} / {proj_title}')
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='项目ID').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=str(project_id))
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='业务月').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=business_month)
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='申请日').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=str(apply_date or '业务月最后一天'))
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='模式').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=('预付' if mode == 'prepay' else '普通'))
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='出款比例').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=profit)
    ws_r.cell(row=ws_r.max_row + 1, column=1, value='天扣除工时').font = TOTAL_FONT
    ws_r.cell(row=ws_r.max_row, column=2, value=daily_ded)
    if calc_err:
        ws_r.cell(row=ws_r.max_row + 1, column=1,
                   value=f'⚠ calc 错误：{calc_err}').font = Font(bold=True, color='C00000')
    # 空一行
    ws_r.append([])
    # 17/13 项表头
    hr = ws_r.max_row + 1
    headers = ['#', '输出项', '金额', '说明']
    for i, h in enumerate(headers, 1):
        c = ws_r.cell(row=hr, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center')
    rows_def = PREPAY_ROWS_LABELS if mode == 'prepay' else NORMAL_ROWS_LABELS
    final_key = '#13_最终出款' if mode == 'prepay' else '#17_最终出款'
    for n, k, label in rows_def:
        amt = detail.get(k)
        desc = detail.get(k + '_说明', '')
        is_pct = isinstance(amt, str) and amt.endswith('%')
        amt_val = amt if (amt is None or is_pct) else (
            float(amt) if isinstance(amt, (int, float)) else amt)
        r_idx = ws_r.max_row + 1
        ws_r.cell(row=r_idx, column=1, value=n)
        ws_r.cell(row=r_idx, column=2, value=label)
        if amt_val is None:
            ws_r.cell(row=r_idx, column=3, value='--')
        else:
            c = ws_r.cell(row=r_idx, column=3, value=amt_val)
            if isinstance(amt_val, (int, float)):
                c.number_format = '#,##0.00'
        ws_r.cell(row=r_idx, column=4, value=desc)
        if k == final_key:
            for ci in range(1, 5):
                cc = ws_r.cell(row=r_idx, column=ci)
                cc.font = TOTAL_FONT; cc.fill = TOTAL_FILL

    # ============ Sheet 1: 考勤预估（按姓名合并） ============
    ws1 = _ws_init(wb, '考勤预估',
                    ['姓名', '工种', '工人类别', '场地/班组', '班次',
                     '出勤天数', '工时合计', '数量合计', '单价', '单位', '估算金额'],
                    [12, 12, 12, 14, 14, 10, 10, 10, 8, 10, 12])
    cur.execute("""SELECT a.name_raw, MIN(a.worker_type), MIN(a.worker_class),
                          MIN(a.floor_or_group), MIN(a.shift_name),
                          COUNT(DISTINCT a.shift_date),
                          SUM(COALESCE(a.hours, 0)), SUM(COALESCE(a.quantity, 0))
                   FROM attendance a
                   WHERE a.project_id=%s AND a.business_month=%s
                   GROUP BY a.name_raw
                   ORDER BY a.name_raw""",
                (project_id, business_month))
    file_cache = {}
    sum_hours = sum_quantity = sum_amount = sum_days = 0
    for n, wt, wc, fl, sn, days, h, q in cur.fetchall():
        price, unit = get_price(sn or '')
        is_day = '天' in (unit or '')
        qty = float(q or 0) if is_day else float(h or 0)
        amt = qty * (price or 0)
        sum_hours += float(h or 0)
        sum_quantity += float(q or 0)
        sum_amount += amt
        sum_days += int(days or 0)
        ws1.append([n or '', wt or '', wc or '', fl or '', sn or '',
                    int(days or 0), float(h or 0), float(q or 0),
                    price or 0, unit, round(amt, 2)])
    # 合计 + 天扣除提示
    _append_total(ws1, 1, '合计',
                   [(7, round(sum_hours, 2)), (8, round(sum_quantity, 2)),
                    (11, round(sum_amount, 2))])
    if daily_ded > 0:
        cur.execute("""SELECT COUNT(DISTINCT shift_date) FROM attendance
                       WHERE project_id=%s AND business_month=%s
                         AND shift_date IS NOT NULL""", (project_id, business_month))
        days = int(cur.fetchone()[0] or 0)
        rep_price = default_price[0] or (next(iter(price_map.values()))[0] if price_map else 0) or 0
        ded_amt = float(daily_ded or 0) * days * float(rep_price or 0)
        r = ws1.max_row + 1
        ws1.cell(row=r, column=1, value=f'天扣除规则: {daily_ded}小时×{days}天×¥{rep_price}').font = TOTAL_FONT
        ws1.cell(row=r, column=11, value=-round(ded_amt, 2)).font = TOTAL_FONT
        ws1.cell(row=r, column=11).number_format = '#,##0.00'
        r2 = ws1.max_row + 1
        ws1.cell(row=r2, column=1, value='扣除后估算').font = TOTAL_FONT
        ws1.cell(row=r2, column=1).fill = TOTAL_FILL
        c = ws1.cell(row=r2, column=11, value=round(max(0, sum_amount - ded_amt), 2))
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL
        c.number_format = '#,##0.00'
    _set_money_col(ws1, 9)
    _set_money_col(ws1, 11)

    # ============ Sheet 2: 账单金额 ============
    ws2 = _ws_init(wb, '账单金额',
                    ['维度', '业务月', '姓名', '身份证号', '金额', '来源类型', '来源文件'],
                    [12, 10, 12, 22, 12, 16, 38])
    sum_bt = 0
    cur.execute("""SELECT business_month, amount, source_type, source_file_id
                   FROM bill_totals WHERE project_id=%s AND business_month=%s
                   ORDER BY id""", (project_id, business_month))
    for bm, amt, st, fid in cur.fetchall():
        if fid not in file_cache: file_cache[fid] = _file_label(cur, fid)
        ws2.append(['合计行', bm, '', '', float(amt or 0), st or '', file_cache[fid]])
        sum_bt += float(amt or 0)

    sum_bp = 0
    cur.execute("""SELECT business_month, name_raw, amount, source_type, source_file_id
                   FROM bill_persons WHERE project_id=%s AND business_month=%s
                   ORDER BY name_raw""", (project_id, business_month))
    for bm, n, amt, st, fid in cur.fetchall():
        if fid not in file_cache: file_cache[fid] = _file_label(cur, fid)
        ws2.append(['人员行', bm, n or '', '', float(amt or 0), st or '', file_cache[fid]])
        sum_bp += float(amt or 0)

    _append_total(ws2, 1, '合计行汇总', [(5, round(sum_bt, 2))])
    _append_total(ws2, 1, '人员行汇总', [(5, round(sum_bp, 2))])
    _set_money_col(ws2, 5)

    # ============ Sheet 3: 发薪流水 ============
    ws3 = _ws_init(wb, '发薪流水',
                    ['付款时间', '业务班次日', '姓名', '身份证号', '金额', '类型', '状态', '来源文件', '备注'],
                    [18, 12, 12, 22, 12, 14, 10, 38, 28])
    sum_pay = 0
    cur.execute("""SELECT pay_time, parsed_shift_date, name_raw, id_card_raw,
                          work_amount, payroll_kind, alipay_status, source_file_id,
                          source_ref
                   FROM payrolls WHERE project_id=%s AND business_month=%s
                   ORDER BY pay_time""", (project_id, business_month))
    for pt, sd, n, ic, wa, pk, st, fid, ref in cur.fetchall():
        if fid not in file_cache: file_cache[fid] = _file_label(cur, fid)
        ws3.append([pt, sd, n or '', ic or '', float(wa or 0),
                    pk or '', st or '', file_cache[fid], (ref or '')[:80]])
        sum_pay += float(wa or 0)
    _append_total(ws3, 1, '合计', [(5, round(sum_pay, 2))])
    _set_money_col(ws3, 5)

    # ============ Sheet 4: 工资表 ============
    ws4 = _ws_init(wb, '工资表',
                    ['业务月', '姓名', '身份证号', '应发工资', '是否代发', '代发对象',
                     '来源文件', '来源行'],
                    [10, 12, 22, 12, 10, 12, 38, 16])
    sum_ws_payable = 0
    cur.execute("""SELECT w.business_month, w.name_raw, '', w.payable_amount,
                          w.is_substitute, w.substitute_name,
                          w.source_file_id, w.source_ref
                   FROM wage_sheets w
                   WHERE w.project_id=%s AND w.business_month=%s
                   ORDER BY w.name_raw""", (project_id, business_month))
    for bm, n, ic, pa, sub, sn, fid, ref in cur.fetchall():
        if fid not in file_cache: file_cache[fid] = _file_label(cur, fid)
        ws4.append([bm, n or '', ic or '', float(pa or 0),
                    '是' if sub else '否', sn or '',
                    file_cache[fid], (ref or '')[:40]])
        sum_ws_payable += float(pa or 0)
    _append_total(ws4, 1, '合计', [(4, round(sum_ws_payable, 2))])
    _set_money_col(ws4, 4)

    # ============ Sheet 5: 考勤×发薪（按姓名合并） ============
    cur.execute("""SELECT name_raw, SUM(hours), SUM(quantity), MIN(shift_name)
                   FROM attendance
                   WHERE project_id=%s AND business_month=%s
                   GROUP BY name_raw""",
                (project_id, business_month))
    att_map = {}
    for n, h, q, sn in cur.fetchall():
        att_map[n] = {'hours': float(h or 0), 'quantity': float(q or 0), 'shift': sn}

    cur.execute("""SELECT name_raw, SUM(work_amount), COUNT(*)
                   FROM payrolls
                   WHERE project_id=%s AND business_month=%s
                   GROUP BY name_raw""",
                (project_id, business_month))
    pay_map = {}
    for n, wa, cnt in cur.fetchall():
        pay_map[n] = {'amount': float(wa or 0), 'count': int(cnt or 0)}

    ws5 = _ws_init(wb, '考勤×发薪',
                    ['姓名', '考勤工时', '考勤数量(天)', '考勤估算金额',
                     '发薪笔数', '发薪合计', '差额(估算-发薪)', '名单'],
                    [12, 10, 12, 14, 10, 12, 14, 10])
    keys = sorted(set(att_map) | set(pay_map))
    sum_a_h = sum_a_q = sum_a_amt = sum_p_amt = 0
    for n in keys:
        a = att_map.get(n)
        p = pay_map.get(n)
        h = a['hours'] if a else 0
        q = a['quantity'] if a else 0
        sn = a['shift'] if a else ''
        price, unit = get_price(sn or '')
        is_day = '天' in (unit or '')
        a_amt = (q if is_day else h) * (price or 0)
        p_amt = p['amount'] if p else 0
        p_cnt = p['count'] if p else 0
        if a and p: status = '都在'
        elif a: status = '考勤'
        else: status = '发薪'
        ws5.append([n or '', round(h, 2), round(q, 2), round(a_amt, 2),
                    p_cnt, round(p_amt, 2), round(a_amt - p_amt, 2), status])
        sum_a_h += h; sum_a_q += q; sum_a_amt += a_amt; sum_p_amt += p_amt
    _append_total(ws5, 1, '合计',
                   [(2, round(sum_a_h, 2)), (3, round(sum_a_q, 2)),
                    (4, round(sum_a_amt, 2)), (6, round(sum_p_amt, 2)),
                    (7, round(sum_a_amt - sum_p_amt, 2))])
    for ci in (4, 6, 7):
        _set_money_col(ws5, ci)

    # ============ Sheet 6: 考勤×工资表 ============
    cur.execute("""SELECT name_raw, SUM(payable_amount)
                   FROM wage_sheets
                   WHERE project_id=%s AND business_month=%s
                   GROUP BY name_raw""",
                (project_id, business_month))
    wage_map = {}
    for n, pa in cur.fetchall():
        wage_map[n] = float(pa or 0)

    ws6 = _ws_init(wb, '考勤×工资表',
                    ['姓名', '考勤工时', '考勤数量(天)', '考勤估算金额',
                     '工资表应发', '差额(估算-应发)', '名单'],
                    [12, 10, 12, 14, 12, 14, 10])
    keys2 = sorted(set(att_map) | set(wage_map))
    s_h = s_q = s_a_amt = s_w = 0
    for n in keys2:
        a = att_map.get(n)
        h = a['hours'] if a else 0
        q = a['quantity'] if a else 0
        sn = a['shift'] if a else ''
        price, unit = get_price(sn or '')
        is_day = '天' in (unit or '')
        a_amt = (q if is_day else h) * (price or 0)
        w_amt = wage_map.get(n, 0)
        in_w = n in wage_map
        if a and in_w: status = '都在'
        elif a: status = '考勤'
        else: status = '工资表'
        ws6.append([n or '', round(h, 2), round(q, 2), round(a_amt, 2),
                    round(w_amt, 2), round(a_amt - w_amt, 2), status])
        s_h += h; s_q += q; s_a_amt += a_amt; s_w += w_amt
    _append_total(ws6, 1, '合计',
                   [(2, round(s_h, 2)), (3, round(s_q, 2)),
                    (4, round(s_a_amt, 2)), (5, round(s_w, 2)),
                    (6, round(s_a_amt - s_w, 2))])
    for ci in (4, 5, 6):
        _set_money_col(ws6, ci)

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f'{ent_short}_{proj_title}_{business_month}_明细.xlsx'
