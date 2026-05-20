import os
import re
import openpyxl
from config import BASE_DATA_DIR

# 排除的目录
EXCLUDE_DIRS = {'1企业模版'}

# 内存缓存
_cache = {}


def clear_cache():
    """清空所有缓存"""
    _cache.clear()


def list_companies():
    """扫描企业目录，返回公司列表"""
    if 'companies' in _cache:
        return _cache['companies']

    enterprise_dir = os.path.join(BASE_DATA_DIR, "企业")
    if not os.path.isdir(enterprise_dir):
        return []

    companies = []
    for name in sorted(os.listdir(enterprise_dir)):
        if name in EXCLUDE_DIRS:
            continue
        full_path = os.path.join(enterprise_dir, name)
        if os.path.isdir(full_path):
            companies.append(name)

    _cache['companies'] = companies
    return companies


def list_projects(company):
    """扫描指定公司下的项目目录"""
    cache_key = f'projects_{company}'
    if cache_key in _cache:
        return _cache[cache_key]

    project_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目")
    if not os.path.isdir(project_dir):
        return []

    projects = []
    for name in sorted(os.listdir(project_dir)):
        full_path = os.path.join(project_dir, name)
        if os.path.isdir(full_path):
            projects.append(name)

    _cache[cache_key] = projects
    return projects


def list_months(count=12, company=None, project=None):
    """返回最近N个月对象列表 [{'key','label','start','end'}, ...]，最新在前。
    若给了 company+project，按项目业务周期返回带区间 label。
    跨月业务周期下：进入下个业务周期时（如澳思美 26-25，今天 day >= 26）下拉首项即下个业务周期。"""
    from datetime import date
    biz = _load_biz_cycle(company, project) if (company and project) else None
    today = date.today()

    # 确定"当前所属业务周期"的月份号
    if biz and biz[2]:
        start_day, _, _ = biz
        if today.day >= start_day:
            # 已进入下个业务周期
            y, m = (today.year, today.month + 1) if today.month < 12 else (today.year + 1, 1)
        else:
            y, m = today.year, today.month
    else:
        y, m = today.year, today.month

    items = []
    for _ in range(count):
        items.append(_month_obj(y, m, biz))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return items


def _parse_month_key(name):
    """将 '2026年3月' 转为 (2026, 3) 用于排序"""
    m = re.match(r'(\d{4})年(\d{1,2})月', name)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return (0, 0)


def _parse_biz_cycle(cycle_str):
    """解析 '上月26-本月25' → (start_day, end_day, cross_month)。否则按自然月。"""
    m = re.match(r'上月(\d+)-本月(\d+)', str(cycle_str or ''))
    if m:
        return (int(m.group(1)), int(m.group(2)), True)
    return (1, 31, False)


def _month_obj(year, month, biz_cycle):
    """返回 {key, label, start, end}。biz_cycle = (start_day, end_day, cross) 或 None（自然月）"""
    from datetime import date
    import calendar
    key = f'{year}年{month}月'
    if biz_cycle and biz_cycle[2]:
        start_day, end_day, _ = biz_cycle
        prev_y = year if month > 1 else year - 1
        prev_m = month - 1 if month > 1 else 12
        s = date(prev_y, prev_m, start_day)
        e = date(year, month, end_day)
        label = f'{key} ({s.month}.{s.day}-{e.month}.{e.day})'
    else:
        last = calendar.monthrange(year, month)[1]
        s = date(year, month, 1)
        e = date(year, month, last)
        label = key
    return {'key': key, 'label': label, 'start': s.isoformat(), 'end': e.isoformat()}


def _load_biz_cycle(company, project):
    """从项目信息.xlsx 基本信息 sheet 读取 业务周期 字段，返回 (start_day, end_day, cross)"""
    info_file = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "项目信息.xlsx")
    if not os.path.isfile(info_file):
        return _parse_biz_cycle(None)
    try:
        wb = openpyxl.load_workbook(info_file, data_only=True, read_only=True)
        if '基本信息' not in wb.sheetnames:
            wb.close()
            return _parse_biz_cycle(None)
        ws = wb['基本信息']
        cycle = None
        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 2 and row[0] and '业务周期' in str(row[0]):
                cycle = row[1]
                break
        wb.close()
        return _parse_biz_cycle(cycle)
    except Exception:
        return _parse_biz_cycle(None)


def _find_latest_month(directory, file_filter=None):
    """扫描目录下匹配的xlsx文件，返回最新月份的sheet名"""
    if not os.path.isdir(directory):
        return None

    latest = None
    latest_key = (0, 0)

    for fname in os.listdir(directory):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        if file_filter and not file_filter(fname):
            continue
        fpath = os.path.join(directory, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            for sheet_name in wb.sheetnames:
                key = _parse_month_key(sheet_name)
                if key > latest_key:
                    latest_key = key
                    latest = sheet_name
            wb.close()
        except Exception:
            continue

    return latest


def _find_latest_date_in_period(directory, file_filter, biz, month_key):
    """扫描目录下匹配的xlsx文件，返回选中业务周期内的最大日期 (YYYY-MM-DD) 或 None。
    自动从行内"日期/班次日期/考勤日期/付款时间"列识别；从 shift_title (如 "3月25白") 兜底。"""
    if not os.path.isdir(directory) or not month_key:
        return None
    from datetime import datetime as _dt
    y, m = _parse_month_key(month_key)
    if y == 0: return None
    obj = _month_obj(y, m, biz)
    p_start = _dt.fromisoformat(obj['start']).date()
    p_end = _dt.fromisoformat(obj['end']).date()

    DATE_HEADERS = ('日期', '班次日期', '考勤日期', '付款时间')
    max_dt = None
    has_period_sheet = False  # 是否有 sheet 名匹配本业务月（兜底用）
    for fname in os.listdir(directory):
        if not fname.endswith('.xlsx') or fname.startswith('~$'): continue
        if file_filter and not file_filter(fname): continue
        fpath = os.path.join(directory, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                # sheet 名匹配业务月 sheet（如 '2026年3月'）→ 标记有数据
                if sn == month_key:
                    has_period_sheet = True
                ws = wb[sn]
                hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not hdr_row: continue
                date_idxs = [i for i, h in enumerate(hdr_row) if h and any(k == str(h) or k in str(h) for k in DATE_HEADERS)]
                if not date_idxs: continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for di in date_idxs:
                        if di >= len(row) or row[di] is None: continue
                        ds = str(row[di])[:10]
                        try:
                            d = _dt.strptime(ds, '%Y-%m-%d').date()
                        except:
                            try: d = _dt.fromisoformat(ds).date()
                            except: continue
                        if d < p_start or d > p_end: continue
                        if max_dt is None or d > max_dt:
                            max_dt = d
            wb.close()
        except Exception:
            continue
    if max_dt:
        return max_dt.isoformat()
    # 兜底：行内无日期但有匹配业务月 sheet（账单/工资表月级汇总）→ 取业务月末与今天的较小者
    if has_period_sheet:
        from datetime import date as _date_today
        today = _date_today.today()
        cap_dt = min(p_end, today)
        return cap_dt.isoformat()
    return None


def load_data_status(company, project, month=None):
    """加载项目的四项数据最新状态：考勤、账单、发薪、工资表
    给了 month（业务周期 key 如 '2026年5月'）→ 显示该周期内最大日期 (YYYY-MM-DD)
    否则 → 显示月份名（旧行为）
    数据查找顺序：共享 → 项目专属（与 v2 工具一致）"""
    att_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "考勤账单")
    pay_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "发薪工资表")
    shared_pay_dir = os.path.join(BASE_DATA_DIR, "企业", company, "共享", "发薪工资表共享")
    shared_bill_dir = os.path.join(BASE_DATA_DIR, "企业", company, "共享", "考勤账单共享")
    pay_dirs = [shared_pay_dir, pay_dir]   # 共享优先
    att_dirs = [shared_bill_dir, att_dir]

    # 考勤：考勤账单目录下，文件名含"考勤"或不含"账单"和"工资表"（即默认当考勤）
    def is_attendance(f):
        name = f.replace('.xlsx', '')
        return '考勤' in name or ('账单' not in name and '工资表' not in name)

    # 账单：文件名含"账单"
    def is_bill(f):
        name = f.replace('.xlsx', '')
        return '账单' in name

    # 发薪：发薪工资表目录下，文件名含"发薪"或不含"工资表"
    def is_payroll(f):
        name = f.replace('.xlsx', '')
        return '发薪' in name or '工资表' not in name

    # 工资表：发薪工资表目录下文件名含"工资表"，或考勤账单目录下文件名含"工资表"
    def is_wage(f):
        name = f.replace('.xlsx', '')
        return '工资表' in name

    def _try_dirs_period(dirs, ff, biz, m):
        for d in dirs:
            v = _find_latest_date_in_period(d, ff, biz, m)
            if v: return v
        return None

    def _try_dirs_month(dirs, ff):
        for d in dirs:
            v = _find_latest_month(d, ff)
            if v: return v
        return None

    if month:
        biz = _load_biz_cycle(company, project)
        attendance = _try_dirs_period(att_dirs, is_attendance, biz, month)
        bill = _try_dirs_period(att_dirs, is_bill, biz, month)
        payroll = _try_dirs_period(pay_dirs, is_payroll, biz, month)
        wage_pay = _try_dirs_period(pay_dirs, is_wage, biz, month)
        wage_att = _try_dirs_period(att_dirs, is_wage, biz, month)
        wage = max([d for d in (wage_pay, wage_att) if d], default=None)
    else:
        attendance = _try_dirs_month(att_dirs, is_attendance)
        bill = _try_dirs_month(att_dirs, is_bill)
        payroll = _try_dirs_month(pay_dirs, is_payroll)
        wage_pay = _try_dirs_month(pay_dirs, is_wage)
        wage_att = _try_dirs_month(att_dirs, is_wage)
        if wage_pay and wage_att:
            wage = wage_pay if _parse_month_key(wage_pay) >= _parse_month_key(wage_att) else wage_att
        else:
            wage = wage_pay or wage_att

    unit_prices = load_unit_prices(company, project)
    # 发薪覆盖率指标：发薪流水 vs 工资表 (按业务月度 sheet 累加)
    coverage = _compute_payroll_coverage(company, project, month, pay_dirs, att_dirs)
    return {
        "attendance": attendance,
        "bill": bill,
        "payroll": payroll,
        "wage": wage,
        "unit_prices": unit_prices,
        # 向后兼容：单价数组只有 1 行时同时给 unit_price 单值
        "unit_price": ({'price': unit_prices[0]['price'], 'unit': unit_prices[0]['unit']}
                       if len(unit_prices) == 1 else None),
        "months": list_project_months(company, project),
        "coverage": coverage,
    }


def _sum_sheet_amount(file_path, sheet_label, amount_col_keywords=('付款金额','金额','应发工资','结算工资','工资','应发')):
    """累加文件指定 sheet 内"金额"列；找不到 sheet 或列时返回 None"""
    if not file_path or not os.path.isfile(file_path):
        return None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_label not in wb.sheetnames:
            wb.close(); return None
        ws = wb[sheet_label]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return None
    if not rows: return None
    # 找表头行（含'姓名'）
    hdr_idx = next((i for i, r in enumerate(rows[:5]) if r and any(v and '姓名' in str(v) for v in r)), 0)
    hdr = rows[hdr_idx]
    amt_i = None
    for kw in amount_col_keywords:
        for i, v in enumerate(hdr):
            if v and kw in str(v):
                amt_i = i; break
        if amt_i is not None: break
    if amt_i is None: return None
    total = 0; cnt = 0
    for r in rows[hdr_idx+1:]:
        if not r or len(r) <= amt_i: continue
        try: amt = float(r[amt_i]) if r[amt_i] not in (None, '') else 0
        except: continue
        if amt <= 0: continue
        total += amt; cnt += 1
    return {'amount': total, 'rows': cnt}


def _compute_payroll_coverage(company, project, month, pay_dirs, att_dirs):
    """计算指定业务月份的发薪流水/工资表金额 + 覆盖率提示。
    month 形如 'YYYY年M月'（list_project_months 的 key）。
    覆盖率 = 发薪流水金额 / 工资表金额。<50% 时给"建议补工资表走规则1C"提示。"""
    if not month: return None
    payroll_file = None
    for d in pay_dirs:
        if not os.path.isdir(d): continue
        for f in os.listdir(d):
            if f.startswith('~$') or not f.endswith('.xlsx'): continue
            if '发薪' in f and '工资表' not in f:
                payroll_file = os.path.join(d, f); break
        if payroll_file: break
    wage_file = None
    for d in pay_dirs + att_dirs:
        if not os.path.isdir(d): continue
        for f in os.listdir(d):
            if f.startswith('~$') or not f.endswith('.xlsx'): continue
            if '工资表' in f:
                wage_file = os.path.join(d, f); break
        if wage_file: break
    p = _sum_sheet_amount(payroll_file, month) if payroll_file else None
    w = _sum_sheet_amount(wage_file, month) if wage_file else None
    payroll_amt = p['amount'] if p else 0
    wage_amt = w['amount'] if w else 0
    pct = (payroll_amt / wage_amt) if wage_amt > 0 else None
    note = None
    if wage_amt > 0 and pct is not None:
        if pct < 0.5:
            note = f'⚠️ 发薪流水仅覆盖工资表 {pct*100:.0f}%，主体工资疑走线下，建议按工资表口径走规则1C「预计直发」'
        elif pct < 0.8:
            note = f'发薪流水覆盖工资表 {pct*100:.0f}%，工资表可作"预计直发"补充'
    return {
        'month': month,
        'payroll_amount': payroll_amt,
        'wage_amount': wage_amt,
        'coverage_pct': pct,
        'note': note,
    }


def load_unit_prices(company, project):
    """从项目信息.xlsx '单价' sheet 读取最新一组生效的所有维度行。

    返回数组 [{area, gz, price, unit, start, note, label}, ...]
    - 多行（如希锐丽盈：白班/夜班）→ web 渲染多个输入框
    - 单行 → 渲染单个输入框（兼容旧项目）
    - "最新一组"按 start（开始时间）取最大值的所有行（同 start 视为同一组）
    """
    info_file = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "项目信息.xlsx")
    if not os.path.isfile(info_file):
        return []
    try:
        wb = openpyxl.load_workbook(info_file, data_only=True)
        if '单价' not in wb.sheetnames:
            wb.close()
            return []
        ws = wb['单价']
        all_rows = []
        max_start = (0, 0)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            try:
                price = float(row[2]) if row[2] is not None else None
            except (TypeError, ValueError):
                continue
            if price is None or price <= 0:
                continue
            def _norm(v):
                if v is None: return ''
                s = str(v).strip()
                return '' if s in ('', '0') else s
            area = _norm(row[0])
            gz = _norm(row[1] if len(row) > 1 else None)
            start = row[3] if len(row) > 3 else None
            note = str(row[5]) if len(row) > 5 and row[5] else ''
            unit = '元/件' if '件' in note else '元/小时'
            start_key = _parse_month_key(str(start)) if start else (0, 0)
            label_parts = [p for p in [area, gz] if p]
            label = '+'.join(label_parts) if label_parts else '默认'
            all_rows.append({
                'area': area, 'gz': gz, 'price': price, 'unit': unit,
                'start': str(start) if start else '', 'note': note,
                'label': label, '_start_key': start_key,
            })
            if start_key > max_start:
                max_start = start_key
        wb.close()
        # 取最新一组：start 等于 max_start 的所有行
        if not all_rows: return []
        latest_group = [r for r in all_rows if r['_start_key'] == max_start]
        for r in latest_group: r.pop('_start_key', None)
        return latest_group
    except Exception:
        return []


def save_unit_prices(company, project, unit_prices):
    """把 web 提交的单价数组写回项目信息.xlsx 单价 sheet。
    按 (area, gz) 匹配最新一组（max start）行，覆盖单价；找不到的新增行。"""
    info_file = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "项目信息.xlsx")
    if not os.path.isfile(info_file):
        return False
    wb = openpyxl.load_workbook(info_file)
    if '单价' not in wb.sheetnames:
        ws = wb.create_sheet('单价')
        ws.append(['场地', '工种', '单价', '开始时间', '结束时间', '备注'])
    else:
        ws = wb['单价']
    # 找出最新一组的 start_key
    max_start = (0, 0)
    rows_idx = []  # [(row_idx, area, gz, start_key)]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 3:
            rows_idx.append((i, '', '', (0, 0))); continue
        def _norm(v):
            if v is None: return ''
            s = str(v).strip()
            return '' if s in ('', '0') else s
        area = _norm(row[0]); gz = _norm(row[1] if len(row) > 1 else None)
        start = row[3] if len(row) > 3 else None
        sk = _parse_month_key(str(start)) if start else (0, 0)
        rows_idx.append((i, area, gz, sk))
        if sk > max_start: max_start = sk
    # 按 (area, gz) 匹配 max_start 组中的行 → 覆盖单价；找不到则 append
    latest_match = {(a, g): idx for idx, a, g, sk in rows_idx if sk == max_start}
    for up in unit_prices:
        ua = (up.get('area') or '').strip()
        ug = (up.get('gz') or '').strip()
        try: pr = float(up.get('price'))
        except (TypeError, ValueError): continue
        idx = latest_match.get((ua, ug))
        if idx:
            ws.cell(row=idx, column=3).value = pr
        else:
            # 新增行：复用 max_start 对应"开始时间"
            start_str = ''
            for ri, a, g, sk in rows_idx:
                if sk == max_start:
                    start_str = ws.cell(row=ri, column=4).value or ''
                    break
            ws.append([ua, ug, pr, start_str, '', 'web 提交新增'])
    wb.save(info_file)
    return True


def list_project_months(company, project):
    """扫描项目下所有 xlsx，收集出现过的月份 sheet 名（YYYY年M月）。
    返回对象列表 [{key, label, start, end}, ...]，按业务周期含日期区间 label，倒序。"""
    project_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project)
    months = set()
    for sub in ('考勤账单', '发薪工资表'):
        sub_dir = os.path.join(project_dir, sub)
        if not os.path.isdir(sub_dir):
            continue
        for fname in os.listdir(sub_dir):
            if not fname.endswith('.xlsx') or fname.startswith('~$'):
                continue
            try:
                wb = openpyxl.load_workbook(os.path.join(sub_dir, fname), read_only=True)
                for sn in wb.sheetnames:
                    if re.match(r'\d{4}年\d{1,2}月', sn):
                        months.add(sn)
                wb.close()
            except Exception:
                continue
    biz = _load_biz_cycle(company, project)
    result = []
    for key in sorted(months, key=_parse_month_key, reverse=True):
        y, m = _parse_month_key(key)
        result.append(_month_obj(y, m, biz))
    return result


def load_project_info(company, project):
    """读取项目信息.xlsx，返回预览用的关键字段"""
    info_file = os.path.join(
        BASE_DATA_DIR, "企业", company, "项目", project, "项目信息.xlsx"
    )
    if not os.path.isfile(info_file):
        return {}

    try:
        wb = openpyxl.load_workbook(info_file, read_only=True)
        # 读基本信息sheet（第一个sheet）
        ws = wb.worksheets[0]
        info = {}
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).strip()
                val = str(row[1]).strip()
                info[key] = val
        wb.close()
        return info
    except Exception:
        return {}
