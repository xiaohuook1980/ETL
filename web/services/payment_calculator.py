"""出款计算步骤1-5"""
import os
import re
import math
import openpyxl
from config import BASE_DATA_DIR


def find_header_col(ws, keyword):
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val and keyword in str(val):
            return c
    return None


def ceil_thousand(val):
    return math.ceil(val / 1000) * 1000


def _find_xlsx_files(directory, name_filter=None):
    """扫描目录下的xlsx文件"""
    if not os.path.isdir(directory):
        return []
    files = []
    for f in os.listdir(directory):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            if name_filter is None or name_filter(f):
                files.append(os.path.join(directory, f))
    return files


def load_attendance_data(company, project, year, month):
    """读取考勤数据：人数和工时"""
    att_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "考勤账单")
    sn = f"{year}年{month}月"

    # 考勤文件：排除纯"账单"和"工资表"
    def is_attendance(f):
        name = f.replace('.xlsx', '')
        return '考勤' in name or ('账单' not in name and '工资表' not in name)

    files = _find_xlsx_files(att_dir, is_attendance)
    if not files:
        # fallback: 所有文件
        files = _find_xlsx_files(att_dir)

    total_amount = 0
    person_count = 0
    names = set()

    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            if sn not in wb.sheetnames:
                wb.close()
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                continue

            headers = [str(h).strip() if h else '' for h in rows[0]]

            # 找姓名列和金额列
            name_idx = None
            amount_idx = None
            hours_idx = None
            for i, h in enumerate(headers):
                if '姓名' in h or '名字' in h:
                    name_idx = i
                if '金额' in h or '应发' in h or '合计' in h:
                    amount_idx = i
                if '工时' in h:
                    hours_idx = i

            for row in rows[1:]:
                if name_idx is not None and row[name_idx]:
                    n = str(row[name_idx]).strip()
                    if n:
                        names.add(n)

            # 尝试读金额
            if amount_idx is not None:
                for row in rows[1:]:
                    if row[amount_idx] is not None:
                        try:
                            total_amount += float(row[amount_idx])
                        except (TypeError, ValueError):
                            pass

        except Exception:
            continue

    person_count = len(names)
    return {
        "人数": person_count,
        "金额": total_amount if total_amount > 0 else None,
        "names": names,
    }


def load_bill_data(company, project, year, month):
    """读取账单数据"""
    att_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "考勤账单")
    sn = f"{year}年{month}月"

    def is_bill(f):
        return '账单' in f.replace('.xlsx', '')

    files = _find_xlsx_files(att_dir, is_bill)
    if not files:
        return None

    total = 0
    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            if sn not in wb.sheetnames:
                wb.close()
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                continue

            headers = [str(h).strip() if h else '' for h in rows[0]]
            amount_idx = None
            for i, h in enumerate(headers):
                if '金额' in h or '账单' in h or '合计' in h:
                    amount_idx = i

            if amount_idx is not None:
                for row in rows[1:]:
                    if row[amount_idx] is not None:
                        try:
                            total += float(row[amount_idx])
                        except (TypeError, ValueError):
                            pass
        except Exception:
            continue

    return total if total > 0 else None


def load_payroll_data(company, project, year, month):
    """读取发薪流水数据"""
    pay_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "发薪工资表")
    sn = f"{year}年{month}月"

    def is_payroll(f):
        name = f.replace('.xlsx', '')
        return '发薪' in name or '工资表' not in name

    files = _find_xlsx_files(pay_dir, is_payroll)
    if not files:
        return None, None

    total = 0
    direct_total = 0
    person_count = 0
    names = set()

    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            if sn not in wb.sheetnames:
                wb.close()
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                continue

            headers = [str(h).strip() if h else '' for h in rows[0]]
            name_idx = None
            amount_idx = None
            work_amount_idx = None

            for i, h in enumerate(headers):
                if '姓名' in h or '名字' in h:
                    name_idx = i
                if '付款金额' in h or '金额' in h:
                    amount_idx = i
                if '工资金额' in h:
                    work_amount_idx = i

            # 直发金额优先用工资金额(work_amount)
            use_idx = work_amount_idx if work_amount_idx is not None else amount_idx

            for row in rows[1:]:
                if name_idx is not None and row[name_idx]:
                    names.add(str(row[name_idx]).strip())
                if use_idx is not None and row[use_idx] is not None:
                    try:
                        val = float(row[use_idx])
                        total += val
                        direct_total += val
                    except (TypeError, ValueError):
                        pass

            person_count = len(names)
        except Exception:
            continue

    return direct_total if direct_total > 0 else None, names


def load_wage_table_data(company, project, year, month):
    """读取工资表数据"""
    pay_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "发薪工资表")
    att_dir = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "考勤账单")
    sn = f"{year}年{month}月"

    def is_wage(f):
        return '工资表' in f.replace('.xlsx', '')

    files = _find_xlsx_files(pay_dir, is_wage) + _find_xlsx_files(att_dir, is_wage)
    if not files:
        return None

    total = 0
    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            if sn not in wb.sheetnames:
                wb.close()
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                continue

            headers = [str(h).strip() if h else '' for h in rows[0]]
            amount_idx = None
            for i, h in enumerate(headers):
                if '金额' in h or '总工资' in h or '合计' in h:
                    amount_idx = i

            if amount_idx is not None:
                for row in rows[1:]:
                    if row[amount_idx] is not None:
                        try:
                            total += float(row[amount_idx])
                        except (TypeError, ValueError):
                            pass
        except Exception:
            continue

    return total if total > 0 else None


def load_already_paid(company, project, year, month):
    """读取本项目本账期已垫付金额"""
    mgr_file = os.path.join(BASE_DATA_DIR, "企业", company, "项目", project, "出款管理", "出款管理表.xlsx")
    if not os.path.isfile(mgr_file):
        return 0

    sn = f"{year}年{month}月"
    service_prefix = f"{year}{month:02d}"
    total = 0

    try:
        wb = openpyxl.load_workbook(mgr_file, data_only=True)
        if "出款记录" not in wb.sheetnames:
            wb.close()
            return 0
        ws = wb["出款记录"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            period = str(row[1]) if row[1] else ""
            amount = row[3]
            # 服务周期匹配当月
            if service_prefix in period and amount:
                try:
                    total += float(amount)
                except Exception:
                    pass
        wb.close()
    except Exception:
        pass

    return total


def calculate_payment(company, project, year, month, use_attendance, use_bill, use_payroll, use_wage):
    """执行出款计算，返回关键指标和计算明细"""
    result = {
        "metrics": {},
        "steps": [],
        "final_amount": 0,
    }

    # 先加载所有原始数据
    attendance_data = load_attendance_data(company, project, year, month) if use_attendance else {"人数": 0, "金额": None, "names": set()}
    bill_amount = load_bill_data(company, project, year, month) if use_bill else None
    payroll_total, payroll_names = (load_payroll_data(company, project, year, month) if use_payroll else (None, None))
    wage_total = load_wage_table_data(company, project, year, month) if use_wage else None

    att_amount = attendance_data.get("金额")  # 考勤有金额列时的值
    direct_pay = payroll_total or 0

    # 匹配率
    att_names = attendance_data.get("names", set())
    if att_names and payroll_names:
        matched = att_names & payroll_names
        match_rate = len(matched) / len(att_names) if att_names else 0
        result["metrics"]["发薪人员匹配率"] = f"{match_rate:.2%}"

    # 步骤1: 预计账单金额（优先级：账单 > 考勤金额 > 发薪/0.8估算）
    if bill_amount and bill_amount > 0:
        estimated_bill = bill_amount
        bill_source = "账单"
    elif att_amount and att_amount > 0:
        estimated_bill = att_amount
        bill_source = "考勤金额"
    elif direct_pay > 0:
        # 无账单无考勤金额时，用发薪/0.8作为估算
        estimated_bill = direct_pay / 0.8
        bill_source = "发薪估算"
    else:
        estimated_bill = 0
        bill_source = "无数据"

    result["steps"].append({"步骤": "步骤1-预计账单金额", "值": estimated_bill, "来源": bill_source})
    result["metrics"]["考勤预计"] = estimated_bill

    if estimated_bill <= 0:
        result["error"] = "无法计算预计账单金额（无考勤、账单和发薪数据）"
        return result

    bill_80 = estimated_bill * 0.8
    result["metrics"]["考勤80%"] = bill_80

    # 步骤2: 已直发 + 预计直发
    predicted_direct = 0
    if wage_total and wage_total > 0:
        remaining = max(0, estimated_bill - direct_pay)
        predicted_direct = min(wage_total, remaining)

    result["metrics"]["已发薪金额"] = direct_pay
    result["metrics"]["直发金额"] = direct_pay
    result["steps"].append({"步骤": "步骤2-已直发+预计直发", "已直发": direct_pay, "预计直发": predicted_direct})

    # 步骤3: 出款金额 = min((已直发+预计直发)/0.8, 预计账单×0.8) - 已垫付
    total_direct = direct_pay + predicted_direct
    v1 = total_direct / 0.8 if total_direct > 0 else 0
    v2 = estimated_bill * 0.8

    already_paid = load_already_paid(company, project, year, month)
    raw_amount = max(0, min(v1, v2) - already_paid)
    raw_amount = ceil_thousand(raw_amount)

    result["metrics"]["本期已垫付"] = already_paid
    result["steps"].append({
        "步骤": "步骤3-出款金额",
        "直发/0.8": v1,
        "账单×0.8": v2,
        "已垫付": already_paid,
        "出款金额": raw_amount,
    })

    # 垫付比例和直发比例
    if estimated_bill > 0:
        total_paid = already_paid + raw_amount
        result["metrics"]["垫付比例"] = f"{total_paid / estimated_bill:.2%}"
    if estimated_bill > 0 and direct_pay > 0:
        result["metrics"]["直发比例"] = f"{direct_pay / estimated_bill:.2%}"

    result["final_amount"] = raw_amount
    return result
